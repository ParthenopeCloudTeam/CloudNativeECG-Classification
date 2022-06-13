from flask import Flask, render_template, request, redirect, url_for, Response
from google.oauth2 import service_account 
from google.cloud import pubsub, storage
from openpyxl import Workbook
from openpyxl.writer.excel import save_virtual_workbook
import numpy as np 
import scipy.io as sio
import matplotlib.pyplot as plt 
import matplotlib.ticker as ticker
import io
import pickle
import base64


app = Flask(__name__)
app.config['SECRET_KEY'] = 'h%#mZVa,U~R35[P4'


@app.route("/", methods=['GET', 'POST'])
def index():   
    if request.method == 'POST':
        tax_code = request.form['tax_code']
        file = request.files['file']
        
        ecg = sio.loadmat(file.stream)['val'].squeeze()
        package = {
            "tax_code": tax_code,
            "filename": file.filename,
            "ecg": ecg
        }
        
        pickle_package = pickle.dumps(package)
        
        future = publisher.publish(topic, pickle_package)
        future.result()
        return redirect(url_for('evaluation_page', tax_code=tax_code))
    
    return render_template('index.html')


@app.route("/<tax_code>/evaluation", methods=['GET'])
def evaluation_page(tax_code):
    blob = bucket.get_blob(tax_code)
    print("blob: ", blob)
    if blob is not None:
        
        print('Esiste il file ', tax_code)
        
        ecg_string = blob.download_as_string().decode('utf-8')
        
        labels = blob.metadata["labels"]
        filename = blob.metadata["filename"]
        labels_list = labels.split(" ")
        
        ecg = np.fromstring(ecg_string[1:-1], dtype=np.int16, sep=' ')
        plot_image = generate_plot(ecg, labels_list, filename)
        
        return render_template("result.html", img_data=plot_image.decode('utf-8'), tax_code=tax_code)
    else:
        print('File non esistente')
        return render_template("waiting.html")


def generate_plot(ecg, labels_list, filename):
    fig = plt.figure()
    axis = fig.add_subplot(1, 1, 1)
    axis.plot(ecg, linewidth=1)
    plt.grid(color='pink', linestyle='-.', linewidth=1)
    axis.set_title("ECG: " + filename)
    
    if not "Rumore" in labels_list:
        axis.set_xticklabels(labels_list, rotation=55)
        axis.xaxis.set_major_locator(ticker.LinearLocator(15))
        plt.gcf().subplots_adjust(bottom=0.25)
    else:
        plt.tick_params(axis='x', which='both', bottom=False, top=False, labelbottom=False)
        axis.set_xlabel('Your ECG signal is noisy.', fontsize=12)
        plt.gcf().subplots_adjust(bottom=0.1)
    output = io.BytesIO()
    plt.savefig(output)
    image = base64.b64encode(output.getvalue())
    
    return image


@app.route('/<tax_code>/download', methods=['GET'])
def download_blob(tax_code):
    blob = bucket.get_blob(tax_code)
    if blob.exists():
        ecg_string = blob.download_as_string().decode('utf-8')
        
        wb = Workbook()
        ws = wb.active
        ws['A1'] = "Filename"
        ws['A2'] = "ECG"
        ws['A3'] = "Label"
        ws['B1'] = blob.metadata["filename"]
        ws['B2'] = ecg_string
        ws['B3'] = blob.metadata["labels"]
        
        output = io.BytesIO(save_virtual_workbook(wb))
        filename = tax_code+".xlsx"
        
        return Response(
                output,
                mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={"Content-disposition": "attachment; filename={}".format(filename)},
                )


def initialization():
    global topic, subscription, response
    global publisher, subscriber, bucket
    
    print('Initializing M1.')

    topic = "projects/PROJECT_NAME/topics/ecg"
 
    credentials = service_account.Credentials.from_service_account_file("key.json")
    publisher = pubsub.PublisherClient(credentials=credentials)
   
    bucket_name = "BUCKET_NAME"
    storage_client = storage.Client(credentials=credentials)
    bucket = storage_client.get_bucket(bucket_name)

    np.set_printoptions(threshold=np.inf)


initialization()
